"""Script defining the server app, inherited from the base App class."""
from PyQt5.QtCore import pyqtSignal
from openpyxl import Workbook
import numpy as np

from app import App
from database import Database
from serverSocket_v2 import ServerSocket
from window import Window, Locations

DB_NAME = 'fignos.db'
SQL_FILE = "base_don_v2.sql"


class ServerApp(App):
    """Class defining all methods necessary to the server app."""

    __duration = 10  # In minutes the duration to watch for tendencies
    __refuel_duration = 15  # In minutes the duration of one refuel

    request_refuel = pyqtSignal(str, str, float, bool, bool)

    def __init__(self):
        """Constructor."""

        self.__database = Database(DB_NAME)

        #création d'une table history_champagne
        max_drink = self.__database.select("SELECT MAX(id) FROM drinks")[0][0]
        self.history_champagne = np.zeros((5, max_drink))
        
        self._socket = ServerSocket()

        self.__reset_rooms()

        App.__init__(self)  # Set at the end because it launches the app

    def __reset_rooms(self):
        """Destructor."""
        print("Setting all rooms to disconnected.")
        self.__database.execute("UPDATE rooms SET connected=0")

    def _open_window(self):
        """Method replacing the App's one opening the server window."""
        self._window = Window(self, Locations.RESERVE)

    def _connect_signals(self):
        """Method connecting the signals of the app."""

        # Connecting quit signal
        self._window.quit_app.connect(self.__reset_rooms)

        # Connecting the socket on connection
        self._socket.listening.connect(self.connection_established.emit)

        # Connecting the socket messages
        self._socket.message_received.connect(lambda m: self.decode_message(m[0], client=m[1].peerAddress().toString()))
        
        # Connecting the clients disconnection
        self._socket.client_disconnected.connect(
            lambda x: self.__database.execute("UPDATE rooms SET connected=0 WHERE ip='{}'".format(x)))

        # Connecting the export signal
        self._window.export_db.connect(self.__exportDB)

        # Connecting the drop base signal
        self._window.drop_base.connect(self.__drop_base)

        # Sending the connection infos when requested
        self._window.get_connection_infos.connect(
            lambda: self.send_connection_infos.emit(self._get_connection_infos()))

        # Connecting the socket when requested
        self._window.set_connection_infos.connect(self.__listen)

        # Sending the bar names when requested
        self._window.get_bar_names.connect(lambda: self.send_bar_names.emit(self._get_bar_names()))

        # Sending the champagne names when requested
        self._window.get_champagne_names.connect(lambda: self.send_champagne_names.emit(self._get_champagne_names()))

        # Setting the name when requested
        self._window.set_bar_name.connect(self.__set_name)

        # Connecting the preferences fill
        self._window.ask_preferences.connect(
            lambda: self.fill_preferences.emit((self._socket.get_address().toString(), self._socket.get_port()),
                                               self._name))

        # Connecting the send message signal
        self._window.send_message.connect(lambda x: self.encode_message(action="ME", message=x))

        # Connecting the cancellation of an order
        self._window.cancel_order.connect(lambda x: self.cancel_sale(x))

        # Connecting the sending refuel signal
        self._window.sending_refuel.connect(
            lambda x: self.encode_message(action="RT", drink=x[1], quantity=x[2], recipient=x[0]))

        # Connecting for the connection dialog
        self._window.request_connection_infos.connect(
            lambda: self.open_connection_dialog.emit(self.__get_devices_state()))

    def __listen(self, args):
        """Method handling the socket."""

        try:
            self._socket.listen(*args)
        except Exception as error:
            self._window.open_dialog("Erreur de connection",
                                     "L'erreur suivante est survenue : {}.".format(error),
                                     type="error")

    def __get_devices_state(self):
        """Method returning devices name and their state of connection"""

        return self.__database.select("SELECT name, connected FROM rooms")

    def __exportDB(self, file):
        """Method exporting the data in the DB to an excel file."""

        print('Exporting workbook...', end="")

        wb = Workbook()

        wb.active.title = "Historique"
        for title in ["Stocks", "Boissons", "Salles"]:
            wb.create_sheet(title)

        tables = []
        for table_name in ["history", "stocks", "drinks", "rooms"]:
            columns = [c[1] for c in self.__database.select("PRAGMA table_info({})".format(table_name))]
            content = self.__database.select("SELECT * FROM {}".format(table_name))
            tables.append([columns] + content)

        for sheet, table in zip(wb, tables):
            for i, entry in enumerate(table):
                for j, attribute in enumerate(entry):
                    sheet.cell(row=i + 1, column=j + 1).value = attribute  # Excel is set to begin at 1

        wb.save(file)

        self._window.open_dialog("Export Excel terminé !", "L'export des données sous format Excel s'est achevé.")
        print('Done.')

    def __drop_base(self):
        """Method called to drop the data base."""

        import os

        self.__database.close()

        os.remove(DB_NAME)

        self.__database = Database(DB_NAME)
        self.__database.import_(SQL_FILE)

        self._window.open_dialog("Importation terminée !", "La base a été remise à zéro.")

    def _get_connection_infos(self):
        """Method getting the connection infos (available ips)."""
        return [ip.toString() for ip in self._socket.get_available_ips()]

    def _get_bar_names(self):
        """Method getting the bar names (bars not connected)."""

        resp = self.__database.select("SELECT name FROM rooms WHERE connected=0")
        return [b[0] for b in resp]

    def _get_champagne_names(self):
        """Method getting the champagnes names."""

        resp = self.__database.select("SELECT name FROM drinks WHERE is_champagne=1")
        print("get champ: ",resp)
        return [b[0] for b in resp]

    def __set_name(self, name):
        """Method called to set the name of the server."""
        self._name = name
        self._window.chat_panel.place_name = name

        # Search for a client with the same name in base
        id_ = None
        resp = self.__database.select(
            "SELECT id FROM rooms WHERE name='{}' AND connected=0".format(name))
        if len(resp) > 0:  # Found
            id_ = resp[-1][0]
        else:
            print("No match found for name '{}'".format(name))

        # Modifying data
        if id_:  # Resuming
            self.__database.execute(
                "UPDATE rooms SET ip='{}', connected=1 WHERE id='{}'".format(self._socket.get_address().toString(),
                                                                             id_))
        else:  # New
            self.__database.execute(
                "INSERT INTO rooms (name, ip) VALUES ('{}', '{}')".format(name, self._socket.get_address().toString()))

        self.name_set.emit()

    # def transaction(self, id_drink, id_room, number, is_sale=True, cancel=False):
    #     """Method used to change the quantity of a drink in the table stocks.
    #     If you want to cancel a previous transaction, set cancel to the desired id."""
    #     # resp = self.__database.select(
    #     #     "SELECT quantity FROM stocks WHERE drink={} AND room={}".format(id_drink, id_room))
    #     # if len(resp) == 0:
    #     #     quantity = resp[0][0]
    #     # else:
    #     #     print("Unable to find a unique record for drink '{}' and bar '{}'.".format(id_drink, id_room))
    #     self.__database.execute("UPDATE stocks SET quantity=quantity+{} WHERE drink='{}' AND room='{}'".format(
    #         (-1 ** is_sale) * number, id_drink, id_room))
    #     if not cancel:
    #         self.__database.execute(
    #             "INSERT INTO history(stamp, drink, room, quantity, is_sale) VALUES (datetime('now','+1 hour'), '{}', '{}', {}, {})".format(
    #                 id_drink, id_room, number, int(is_sale)))
    #     # else:
    #     #     self.__notification
    #
    #     # if not is_sale:
    #     #     self.__notification.remove([id_drink, id_room])
    #     self.forecast_all()  # It would be better reforecasting only the drink concerned

    def transaction_v2(self, id_room, id_drink, quantity, sale=True, cancellation=False):
        """Method used to perform a transaction on the database.
        To make a resupply, set sale to False."""

        # Updating the stocks
        self.__database.execute("UPDATE stocks SET quantity=quantity+{} WHERE drink='{}' AND room='{}'".format(
            ((-1) ** sale) * quantity, id_drink, id_room))

        self.__database.execute("UPDATE stocks SET consommation = consommation + {} WHERE drink='{}' AND room='{}'".format(
            sale * quantity , id_drink, id_room))
        
        if cancellation:
            self.__database.execute("UPDATE stocks SET consommation = consommation + {} WHERE drink='{}' AND room='{}'".format(
            -1*quantity , id_drink, id_room))
            
        list_champagne = self.__database.select("SELECT id FROM drinks WHERE is_champagne = 1")
        for i in range(len(list_champagne)):
            list_champagne[i]=list_champagne[i][0]
            if int(id_drink) in list_champagne :
                list_emit= []
                
                self.__database.execute("UPDATE stocks SET quantity_sold=quantity_sold+{} WHERE drink='{}' AND room='{}'".format(
                                        sale * quantity, id_drink, id_room))

                qteSoldChamp = self.__database.select("SELECT s.drink, s.room, s.consommation FROM stocks AS s JOIN drinks AS d WHERE d.id = s.drink AND d.is_champagne = 1")
                
                qteChamp = np.zeros((len(list_champagne),5),int)
                
                for i in qteSoldChamp:
                    qteChamp[i[0]-1][i[1]-1] = i[2]
                
                self.encode_message(action="CH", qteChamp=qteChamp.transpose().tolist())
        
        if not cancellation:
            # Inserting a row in the history
            self.__database.execute(
                "INSERT INTO history(stamp, drink, room, quantity, is_sale) VALUES (datetime('now','+1 hour'), '{}', '{}', {}, {})".format(
                    id_drink, id_room, quantity, int(sale)))

            # updating quantity_sold for the CDF
            if int(id_drink) in list_champagne :
                list_emit= []
                
                self.__database.execute("UPDATE stocks SET quantity_sold=quantity_sold+{} WHERE drink='{}' AND room='{}'".format(
                                        sale * quantity, id_drink, id_room))

                qteSoldChamp = self.__database.select("SELECT s.drink, s.room, s.consommation FROM stocks AS s JOIN drinks AS d WHERE d.id = s.drink AND d.is_champagne = 1")
                
                qteChamp = np.zeros((len(list_champagne),5),int)
                
                for i in qteSoldChamp:
                    qteChamp[i[0]-1][i[1]-1] = i[2]
                
                self.encode_message(action="CH", qteChamp=qteChamp.transpose().tolist())

                ##self.history_champagne[id_room-1, id_drink-1]=(self.__database.select("SELECT LAST_INSERT_ID() FROM history WHERE is_sale"))
                
            

            # Getting the id of the sale
            resp = self.__database.select("SELECT last_insert_rowid()")
            if len(resp) == 1:
                id = resp[0][0]

        # Re forecasting the drink in the room
        time = self.forecast_v2(id_room, id_drink)

        # Fetching the bar name
        bar_name = ""
        resp = self.__database.select("SELECT name, ip FROM rooms WHERE id='{}'".format(id_room))
        if len(resp) == 1:
            bar_name = resp[0][0]

        # Fetching drink name
        drink_name = ""
        resp = self.__database.select("SELECT name FROM drinks WHERE id='{}'".format(id_drink))
        if len(resp) == 1:
            drink_name = resp[0][0]

        # Emitting signal for refuel
        if bar_name and drink_name:  # Infos were found
            self.request_refuel.emit(bar_name, drink_name, time, time <= 2 * self.__refuel_duration,
                                     not sale)  # 2 times the duration of one delivery (it becomes critical)
        else:
            print("Unable to request refuel for parameters ({}, {}, {})".format(bar_name, drink_name, time))

        try:
            print("Bar '{}' can wait for drink '{}' ({}min)".format(id_room, id_drink, int(time)))
        except OverflowError:
            print("Bar '{}' can wait for drink '{}' forever".format(id_room, id_drink))

        # Send confirmation except if it is a cancellation
        if sale and not cancellation:

            drink_name = ""
            bar_name = ""

            # Getting ip and name of the bar
            resp = self.__database.select("SELECT name, ip FROM rooms WHERE id='{}'".format(id_room))
            if len(resp) == 1:
                bar_name = resp[0][0]
                ip = resp[0][1]
            else:
                print("Unable to retrieve the ip of bar '{}' to send a sale confirmation.".format(id_room))

            # Getting the name of the drink
            resp = self.__database.select("SELECT name FROM drinks WHERE id='{}'".format(id_drink))
            if len(resp) == 1:
                drink_name = resp[0][0]
            else:
                print("Unable to retrieve the name of drink '{}' to send a sale confirmation.".format(id_drink))

            # Sending the confirmation
            if drink_name and bar_name:
                self._window.add_order(id, drink_name, quantity, bar_name)
                self.encode_message(action="VE", id=id, drink=id_drink, quantity=quantity, ip=ip)

    def cancel_sale(self, id_order):
        """Method used to cancel an order"""
        print(id_order,type(id_order))
        # Update history table
        self.__database.execute("UPDATE history SET is_cancelled='{}' WHERE id='{}'".format(1, id_order))

        liste = self.__database.select("SELECT drink, room FROM history WHERE id='{}'".format(id_order))
        id_drink, id_room = liste[0]
        
        self.history_champagne[id_room-1, id_drink-1]=0

        # Cancelling sale in stocks
        resp = self.__database.select(
            "SELECT quantity, drink, room FROM history WHERE id='{}'".format(id_order))
        if len(resp) == 1:
            quantity, drink, room = resp[0]

            print("Recrediting the bar '{}' of quantity '{}' of drink '{}'".format(room, quantity, drink))
            self.transaction_v2(room, drink, quantity, False, True)  # Cancelling the sale

            # Getting client's ip
            resp = self.__database.select("SELECT ip FROM rooms WHERE id='{}' AND connected=1".format(room))
            if len(resp) > 0:
                ip = resp[-1][0]
                self._window.remove_cancelled_order(id_order)
                self.encode_message(action="AE", canceled_sale_id=id_order, recipient=ip)
            else:
                print("Unable to cancel sale '{}' because room '{}' ip could not be found".format(id_order, room))
        else:
            print("Unable to retrieve and cancel sale '{}' from history".format(id_order))

    # def get_sale_sum(self, id_drink, id_room, duration=__duration):
    #     "Method used to get the history of orders up until a few minutes"
    #     [[total_sales]] = self.__database.select("SELECT sum(quantity) FROM history \
    #                                               WHERE drink = {} AND room = {} AND is_sale = 1 AND is_cancelled = 0 \
    #                                               AND (stamp >= datetime('now','+1 hour','-{} minutes'))".format(
    #         id_drink, id_room, duration))
    #     # What about last refuelling ?
    #
    #     if type(total_sales) == int:
    #         return total_sales
    #     else:
    #         return 0

    # def forecast(self, id_drink, id_room):
    #     """Method used to forecast the need for a resupply for ONE drink"""
    #     slope = self.get_sale_sum(id_drink, id_room) / self.__duration
    #     [[limit]] = self.__database.select("SELECT treshhold FROM drinks WHERE id = '{}'".format(id_drink))
    #     [[qty]] = self.__database.select(
    #         "SELECT quantity FROM stocks WHERE drink = {} AND room = {}".format(id_drink, id_room))
    #
    #     if slope != 0:
    #         time_appro = (qty - limit) / slope
    #     else:
    #         time_appro = 10000  # This is made just because we need to return
    #         # something from the function
    #
    #     return time_appro, slope

    def forecast_v2(self, id_room, id_drink):
        """Method used to forecast the duration after which the need of refuel will become critical."""

        # Getting the current state of sales
        resp = self.__database.select(
            "SELECT quantity FROM stocks WHERE room='{}' AND drink='{}'".format(id_room, id_drink))
        if len(resp) > 0:
            current = resp[0][0]

            # Getting the total quantity of drinks sold during the duration
            resp = self.__database.select(
                "SELECT sum(quantity) FROM history WHERE drink='{}' AND room='{}' AND is_sale=1 AND is_cancelled=0 AND stamp>=datetime('now', '+1 hour', '-{} minutes')".format(
                    id_drink, id_room, self.__duration))
            if len(resp) == 1 and resp[0][0]:
                quantity = resp[0][0]

                print("A quantity of '{}' of drink '{}' was sold in bar '{}'".format(quantity, id_drink, id_room))

                # Getting the threshold for the drink
                resp = self.__database.select("SELECT threshold from drinks WHERE id='{}'".format(id_drink))
                if len(resp) == 1:
                    threshold = resp[0][0]
                else:
                    threshold = 0

                # Computing the time of shortage
                time = (current - threshold) * self.__duration / quantity  # Time to shortage from now in minutes
                print(
                    "There will be a shortage of drink '{}' in bar '{}' in {} minutes".format(id_drink, id_room, time))
                return time

            else:  # No drink were sold in the last period
                return float("inf")

        else:
            print("The drink '{}' is not sold in the bar '{}'.".format(id_drink, id_room))

    # def forecast_all(self):
    #     """Method used to forecast the need for a resuply for ALL drinks"""
    #     # list_id_room = [element for sublist in self.__database.select("SELECT id FROM rooms") for element in sublist]
    #     list_id_room = [room[0] for room in self.__database.select(
    #         "SELECT id FROM rooms")]  # Ringing a bell ? It is called not being a complete moron
    #     list_id_drink = [element for sublist in self.__database.select("SELECT id FROM drinks") for element in sublist]
    #
    #     # Section to determine a list of restock querries
    #     restock_query_list = []
    #     temporary_query_list = []
    #
    #     for id_room in list_id_room:  # No! No! No! All drinks are not in all rooms!!
    #         for id_drink in list_id_drink:
    #             (time, slope) = self.forecast(id_drink, id_room)
    #             quantity = (
    #                                2 * self.__refuel_duration - time) * slope  # This quantity has been calculated to set the next refuel time at 30 minutes after the query
    #             if [id_drink, id_room] in self.__notification:
    #                 pass  # Do nothing
    #             if time <= self.__refuel_duration:
    #                 restock_query_list.append([id_drink, id_room, quantity])
    #             elif time <= 2 * self._refuel_time:
    #                 temporary_query_list.append([id_drink, id_room, quantity])
    #         if len(restock_query_list) != 0:
    #             restock_query_list += temporary_query_list
    #     return restock_query_list

    # def restock_query(self, query_list):
    #     """Method used to add new notifications for the Zi Bar"""
    #     if len(query_list) == 0:
    #         pass
    #     else:
    #         for query in query_list:
    #             if query[:2] not in self.__notification:
    #                 self.__notification.append(query_list)
    #     # Further uses for notifications ...

    def decode_message(self, message, client=None):
        """Method used to understand received messages."""

        print("Decoding message '{}'".format(message))

        message_split = message[1:-1].split('||')

        if len(message_split) > 1:  # Several messages are queued
            for m in message_split:
                self.decode_message('|' + m + '|', client)
            return
        else:
            message = message_split[0]

        message_split = message.split('|')

        if message_split[0] == "NO":  # selected name

            print("Client '{}' set its name to '{}'".format(client, message_split[1]))

            # Search for a client with the same name in base
            connected = False
            client_id = None
            resp = self.__database.select(
                "SELECT id, connected FROM rooms WHERE name='{}'".format(message_split[1]))
            if len(resp) > 0:  # Client found
                client_id, connected = resp[-1]
            else:
                print("No client disconnected found matching the name '{}'".format(message_split[1]))

            if not connected:  # The client is not already connected somewhere else
                # Modifying data
                if client_id:  # A client is resuming its connection
                    self.__database.execute(
                        "UPDATE rooms SET ip='{}', connected=1 WHERE id='{}'".format(client, client_id))
                else:  # New client
                    self.__database.execute(
                        "INSERT INTO rooms (name, ip) VALUES ('{}', '{}')".format(message_split[1], client))
                    client_id = self.__database.select("SELECT last_inserted_rowid()")

                # Sending the list of drinks
                self.encode_message(action="LO", client_id=client_id, recipient=client)

                # Sending the list of food
                self.encode_message(action="LE", recipient=client)

            else:
                print("A client tried to set its name to one already taken '{}'".format(message_split[1]))

        elif message_split[0] == "VE":  # sale

            # Simpler solution but most efficient would be to compact it
            resp = self.__database.select("SELECT id FROM rooms WHERE ip='{}'AND connected=1".format(client))
            if len(resp) > 0:
                id_bar = resp[-1][0]
                self.transaction_v2(id_bar, message_split[1],
                                    int(message_split[2]))  # Id of the bar, id of the drink and number of drinks
            else:
                print("No bar found matching the given ip '{}'...".format(client))

        elif message_split[0] == "AR":  # client cancel a sale

            print(message_split[1])
            self.cancel_sale(message_split[1])

        elif message_split[0] == "ME":  # Message by a client

            # Getting the client's name
            client_name = ""
            resp = self.__database.select("SELECT name FROM rooms WHERE ip='{}' AND connected=1".format(client))
            if len(resp) > 0:  # Client found
                client_name = resp[-1][0]
                new_message = client_name + '|' + message[3:]
            else:
                new_message = message
                print("No client found matching the ip '{}'".format(client))

            # Checking for the receiver
            dests = [word[1:] for word in message[3:].split(' ') if
                     word.startswith('@')]  # Words starting with '@' in the message
            ips = []
            for d in dests:
                resp = self.__database.select("SELECT ip FROM rooms WHERE name='{}'".format(d.lower()))
                if len(resp) > 0:  # At least one match found
                    ips.append(resp[-1][0])

            # print("Relaying message '{}' to ips '{}'".format(new_message, ips))

            # Preparing the message for the UI
            if client_name:
                infos = (message[3:], client_name)
            else:
                infos = (message[3:],)

            # Sending the message
            if ips:  # Some specific receivers were found

                # Fetching CDF ip
                resp = self.__database.select("SELECT ip FROM rooms WHERE name='cdf'")
                if len(resp) > 0 and resp[-1][0] not in ips:
                    ips.append(resp[-1][0])

                for ip in ips:
                    if ip == self._socket.get_address().toString():
                        self.message_received.emit(infos)  # Sending the message to the UI
                    else:
                        self._socket.send_message("|ME|{}|".format(new_message), ip)
                if client not in ips:
                    self._socket.send_message("|ME|{}|".format(new_message), client)

            else:
                self._socket.send_message("|ME|{}|".format(new_message))
                self.message_received.emit(infos)  # Sending the message to the UI

            # self._socket.send_message(message)  # Broadcasting the message.
            # self.message_received.emit(message_split[1])
            
        elif message_split[0] == "UR":  # Message by a client

            # Getting the client's name
            client_name = ""
            resp = self.__database.select("SELECT name FROM rooms WHERE ip='{}' AND connected=1".format(client))
            if len(resp) > 0:  # Client found
                client_name = resp[-1][0]
                new_message = client_name + '|' + message[3:]
            else:
                new_message = message
                print("No client found matching the ip '{}'".format(client))

            # Checking for the receiver
            dests = [word[1:] for word in message[3:].split(' ') if
                     word.startswith('@')]  # Words starting with '@' in the message
            ips = []
            for d in dests:
                resp = self.__database.select("SELECT ip FROM rooms WHERE name='{}'".format(d.lower()))
                if len(resp) > 0:  # At least one match found
                    ips.append(resp[-1][0])

            # print("Relaying message '{}' to ips '{}'".format(new_message, ips))

            # Preparing the message for the UI
            if client_name:
                infos = (message[3:], client_name)
            else:
                infos = (message[3:],)

            # Sending the message
            if ips:  # Some specific receivers were found

                # Fetching CDF ip
                resp = self.__database.select("SELECT ip FROM rooms WHERE name='cdf'")
                if len(resp) > 0 and resp[-1][0] not in ips:
                    ips.append(resp[-1][0])

                for ip in ips:
                    if ip == self._socket.get_address().toString():
                        self.urgent_message_received.emit(infos)  # Sending the message to the UI
                    else:
                        self._socket.send_message("|UR|{}|".format(new_message), ip)
                if client not in ips:
                    self._socket.send_message("|UR|{}|".format(new_message), client)

            else:
                self._socket.send_message("|UR|{}|".format(new_message))
                self.urgent_message_received.emit(infos)  # Sending the message to the UI

        elif message_split[0] == "RE":  # refueled

            # Getting readable parameters
            _, id_drink, quantity, id_refuel = message_split

            resp = self.__database.select("SELECT id FROM rooms WHERE ip='{}'AND connected=1".format(client))
            if len(resp) > 0:
                id_bar = resp[-1][0]

                # Getting the size of a container
                resp = self.__database.select("SELECT container_size FROM drinks WHERE id='{}'".format(id_drink))
                if len(resp) == 1:
                    container_size = resp[0][0]

                    # We suppose the reserve is never out of stocks
                    self.transaction_v2(id_bar, id_drink, int(quantity) * float(container_size),
                                        sale=False)  # Setting sale to False makes the transaction positive

                    # Notifying the client
                    self.encode_message(action="RE", id_refuel=id_refuel, recipient=client)

                else:
                    print("Unable to retrieve the size of a container of drink '{}' to perform a refuel.".format(
                        id_drink))

            else:
                print("No bar found matching the given ip '{}'...".format(client))

        elif message_split[0] == "RS":  # Restal

            # Fetching restal ip
            resp = self.__database.select("SELECT ip FROM rooms WHERE name='{}'".format('restal'))
            if len(resp) > 0:
                ip = resp[-1][0]

            # Fetching sender bar name
            resp = self.__database.select("SELECT name FROM rooms WHERE ip='{}'".format(client))
            if len(resp) > 0:
                bar = resp[-1][0]

            # Warning the restal about the sale
            self.encode_message(action="RS", bar=bar, food_id=message_split[1], quantity=message_split[2], recipient=ip)

        elif message_split[0] == "LA":  # Requesting a list of bars

            self.encode_message(action='LA', recipient=client)

        elif message_split[0] == "GC":

            self.encode_message(action='GC', recipient=client)
                                                 
        elif message_split[0] == "CH":

            pass

        else:

            self._window.open_dialog("Message incompréhensible",
                                     "Le message suivant n'a pas pu être décodé : {}".format(message), type="warning")
            print("Error : message '%s' could not be decoded" % message)

            

    def encode_message(self, **kwargs):
        """
        :param kwargs: key 'action' related to agreed values: LA - AE - LO - DE - RT - RS - ME - CH
                       with following keys 'action' :
                            LA : use key 'recipient'
                            AE : use key 'canceled_sale_id' and 'recipient'
                            LO : use key 'bar_id' and 'recipient'
                            DE : use key 'recipient'
                            RT : use key 'drink_id'
                            RS : no effect for the moment
                            ME : use key 'message'
                            CH : use key 'qteChamp'
        :return:
        """
        if kwargs["action"] == "LA":  # not connected bars list

            names = self.__database.select("SELECT name FROM rooms WHERE connected=0")
            new_message = "|LA|" + ','.join([str(n[0]) for n in names]) + '|'
            print("Sending list of bars '{}' to client '{}'".format(new_message[3:], kwargs['recipient']))
            self._socket.send_message(new_message, kwargs['recipient'])

        elif kwargs["action"] == "GC":
                
            names = self.__database.select("SELECT name FROM drinks WHERE is_champagne=1")
            rooms = self.__database.select("SELECT name FROM rooms WHERE is_bar=1")
            
            new_message = "|GC|" + ','.join([str(n[0]) for n in names]) + '|' + ','.join([str(n[0]) for n in rooms]) + '|'
            
            self._socket.send_message(new_message, kwargs['recipient'])
                                                 
        elif kwargs["action"] == "AE":  # inform that one sale was canceled

            self._socket.send_message("|%s|%s|" % (kwargs["action"], kwargs["canceled_sale_id"]), kwargs["recipient"])

        elif kwargs["action"] == "LO":  # available drinks list

            # Fetching and sending the list of drinks
            drinks = self.__database.select(
                "SELECT drinks.id, drinks.name, drinks.by_bottle FROM stocks JOIN drinks ON stocks.drink=drinks.id WHERE stocks.room='{}'".format(
                    kwargs['client_id']))
            print("Query returned '{}'".format(drinks))
            print("Sending list message '{}'".format(','.join([str(d[0]) + ':' + str(d[1]) + ':' + str(d[2]) for d in drinks])))
            self._socket.send_message("|LO|" + ','.join([str(d[0]) + ':' + str(d[1]) + ':' + str(d[2]) for d in drinks]) + '|',
                                      kwargs['recipient'])

        elif kwargs['action'] == "LE":  # available food

            # Fetching and sending the list of food
            food = self.__database.select("SELECT id, name FROM food")
            self._socket.send_message("|LE|" + ','.join([str(f[0]) + ':' + str(f[1]) for f in food]) + '|',
                                      kwargs['recipient'])

        elif kwargs["action"] == "RT":  # refuel in progress

            # Getting the id of the drink
            drink = ""
            resp = self.__database.select(
                "SELECT id FROM drinks WHERE name='{}'".format(kwargs['drink']))
            if len(resp) == 1:
                drink = resp[0][0]

            # Getting the ip of the recipient
            ip = ""
            resp = self.__database.select(
                "SELECT ip FROM rooms WHERE name='{}' AND connected=1".format(kwargs['recipient']))
            if len(resp) > 0:
                ip = resp[-1][0]

            if drink and ip:  # Both drink id and ip found
                self._socket.send_message("|RT|{}|{}|".format(drink, kwargs['quantity']), ip)

        elif kwargs["action"] == "RS":  # Restal

            self._socket.send_message("|RS|{}|{}|{}|".format(kwargs['bar'], kwargs['food_id'], kwargs['quantity']),
                                      recipient=kwargs['recipient'])

        elif kwargs["action"] == "ME":  # Message by the server

            print("Encoding message '{}'".format(kwargs["message"]))

            # Getting the server's name
            name = ""
            resp = self.__database.select(
                "SELECT name FROM rooms WHERE ip='{}' AND connected=1".format(self._socket.get_address().toString()))
            if len(resp) > 0:  # At least one match was found
                name = resp[-1][0]  # By default we select the last one
                message = "|ME|" + name + '|' + kwargs["message"] + '|'  # Rewriting the message
            else:
                message = "|ME|" + kwargs["message"] + '|'

            print("Message became '{}'".format(message))

            # Checking for the receiver
            dests = [word[1:] for word in kwargs["message"].split(' ') if
                     word.startswith('@')]  # Words starting with '@' in the message
            ips = []
            for d in dests:
                resp = self.__database.select("SELECT ip FROM rooms WHERE name='{}'".format(d.lower()))
                if len(resp) > 0:  # At least one match found
                    ips.append(resp[-1][0])

            print("Sending message '{}' to ips '{}'.".format(message, ips))

            # Sending the message
            if ips:  # Some specific receivers were found

                # Fetching CDF ip
                resp = self.__database.select("SELECT ip FROM rooms WHERE name='cdf'")
                if len(resp) > 0 and resp[-1][0] not in ips:
                    ips.append(resp[-1][0])

                for ip in ips:
                    self._socket.send_message(message, ip)
                # No need to send the message to himself
            else:
                self._socket.send_message(message)

            # Sending the message to the UI
            if name:
                infos = (kwargs["message"], name)
            else:
                infos = (kwargs["message"],)
            self.message_received.emit(infos)

            # if kwargs["message"][0] == "@":
            #     recipient = kwargs["message"]
            #     for c in kwargs["message"][1:]:
            #         if c != " ":
            #             recipient += c
            #         else:
            #             break  # Not the best practice
            #     # Fetch recipient ip from database
            #     resp = self.__database.select("SELECT ip FROM rooms WHERE name={}".format(recipient.lower()))
            #     if len(resp) == 0:  # No result
            #         print("No room matching recipient '{}'.".format(recipient))
            #         self._socket.send_message("ME|{}".format(kwargs["message"]))
            #     else:  # At least one room was found
            #         ip = resp[-1][0]  # We send the message to the last one
            #         self._socket.send_message("ME|%s" % (kwargs["message"]), recipient=ip)
            # else:  # No recipient provided
            #     self._socket.send_message("ME|{}".format(kwargs["message"]))
            # self.message_received.emit(kwargs["message"])

        elif kwargs["action"] == "UR":  # Urgent-Message by the server

            print("Encoding message '{}'".format(kwargs["message"]))

            # Getting the server's name
            name = ""
            resp = self.__database.select(
                "SELECT name FROM rooms WHERE ip='{}' AND connected=1".format(self._socket.get_address().toString()))
            if len(resp) > 0:  # At least one match was found
                name = resp[-1][0]  # By default we select the last one
                message = "|UR|" + name + '|' + kwargs["message"] + '|'  # Rewriting the message
            else:
                message = "|UR|" + kwargs["message"] + '|'

            print("Message became '{}'".format(message))

            # Checking for the receiver
            dests = [word[1:] for word in kwargs["message"].split(' ') if
                     word.startswith('@')]  # Words starting with '@' in the message
            ips = []
            for d in dests:
                resp = self.__database.select("SELECT ip FROM rooms WHERE name='{}'".format(d.lower()))
                if len(resp) > 0:  # At least one match found
                    ips.append(resp[-1][0])

            print("Sending message '{}' to ips '{}'.".format(message, ips))

            # Sending the message
            if ips:  # Some specific receivers were found

                # Fetching CDF ip
                resp = self.__database.select("SELECT ip FROM rooms WHERE name='cdf'")
                if len(resp) > 0 and resp[-1][0] not in ips:
                    ips.append(resp[-1][0])

                for ip in ips:
                    self._socket.send_message(message, ip)
                # No need to send the message to himself
            else:
                self._socket.send_message(message)

            # Sending the message to the UI
            if name:
                infos = (kwargs["message"], name)
            else:
                infos = (kwargs["message"],)
            self.urgent_message_received.emit(infos)

        elif kwargs["action"] == "VE":  # Confirmation of sales

            try:
                message = "|VE|" + str(kwargs["id"]) + '|' + str(kwargs["drink"]) + '|' + str(kwargs["quantity"]) + '|'
            except KeyError:
                print("Unable to send a confirmation of sales: a parameter is missing.")
            else:
                try:
                    self._socket.send_message(message, kwargs["ip"])
                except KeyError:
                    print("The 'ip' argument must be specified.")

        elif kwargs['action'] == "RE":

            self._socket.send_message("|RE|{}|".format(kwargs['id_refuel']), kwargs['recipient'])

        elif kwargs["action"] == "CH": # Champagne

            self._socket.send_message("|CH|{}|".format(kwargs["qteChamp"]))

        else:

            self._window.open_dialog("Impossible d'envoyer un message",
                                     "Le message suivant n'a pas pu être envoyé car mal encodé : {}".format(kwargs),
                                     type="warning")
            print("Error during encoding \n arguments : %s" % kwargs)
            


if __name__ == '__main__':
    from PyQt5.QtWidgets import QApplication
    from PyQt5.QtCore import *
    import sys

    app = QApplication(sys.argv)
    
    file = QFile("style.qss")
    file.open(QFile.ReadOnly | QFile.Text)
    stream = QTextStream(file)
    app.setStyleSheet(stream.readAll())

    client = ServerApp()
    sys.exit(app.exec_())
